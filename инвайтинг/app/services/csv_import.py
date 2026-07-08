from __future__ import annotations

import csv
import io
import re

from sqlalchemy import insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import InviteTarget
from app.schemas import ImportResponse


_USERNAME_RE = re.compile(r"^[A-Za-z][A-Za-z0-9_]{3,31}$")

_USERNAME_KEYS = {
    "username",
    "user",
    "login",
    "юзернейм",
    "логин",
    "ник",
    "никнейм",
    "аккаунт",
    "telegram",
    "tg",
    "t.me",
}
_USER_ID_KEYS = {
    "user_id",
    "userid",
    "id",
    "tg_id",
    "telegram_id",
    "telegramid",
    "chat_id",
    "айди",
    "ид",
}


def _norm_username(raw: str) -> str | None:
    v = (raw or "").strip()
    if not v:
        return None
    if v.startswith("@"):
        v = v[1:]
    if "t.me/" in v.lower():
        v = v.split("t.me/")[-1]
    v = v.split("?")[0].split("/")[-1].strip()
    if not v or not _USERNAME_RE.match(v):
        return None
    return v


def _parse_user_id(raw: str) -> int | None:
    v = (raw or "").strip()
    if not v:
        return None
    # Excel иногда отдаёт float: 123456.0
    if re.fullmatch(r"\d+\.0+", v):
        v = v.split(".", 1)[0]
    if not v.isdigit():
        return None
    try:
        return int(v)
    except Exception:
        return None


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _pick_from_keyed(keyed: dict[str, str]) -> dict[str, str | None]:
    username = None
    user_id = None
    for k, v in keyed.items():
        key = (k or "").strip().lower()
        if key in _USERNAME_KEYS and v and username is None:
            username = v
        if key in _USER_ID_KEYS and v and user_id is None:
            user_id = v
    return {"username": username, "user_id": user_id}


def _row_from_cells(cells: list[str], has_header: bool, headers: list[str] | None = None) -> dict[str, str | None] | None:
    if has_header and headers:
        keyed = {}
        for i, h in enumerate(headers):
            keyed[h] = cells[i] if i < len(cells) else ""
        return _pick_from_keyed(keyed)

    # без заголовка: первая непустая ячейка
    for cell in cells:
        cell = (cell or "").strip()
        if not cell:
            continue
        if cell.lower() in _USERNAME_KEYS | _USER_ID_KEYS | {"user_id", "username"}:
            return None
        if cell.isdigit() or re.fullmatch(r"\d+\.0+", cell):
            return {"username": None, "user_id": cell}
        return {"username": cell, "user_id": None}
    return None


def _headerish(cells: list[str]) -> bool:
    joined = " ".join((c or "").strip().lower() for c in cells if c)
    if not joined:
        return False
    keys = _USERNAME_KEYS | _USER_ID_KEYS
    return any(k in joined for k in keys)


def _iter_rows_from_text(text: str) -> list[dict[str, str | None]]:
    """
    CSV / TXT:
    - с заголовками username / user_id
    - без заголовка (одна колонка)
    - просто по строкам
    """
    text = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return []

    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except Exception:
        dialect = csv.excel

    stream = io.StringIO(text, newline="")
    first_line = text.split("\n", 1)[0].lower()
    has_header = _headerish([first_line])

    rows: list[dict[str, str | None]] = []
    if has_header:
        reader = csv.DictReader(stream, dialect=dialect)
        for raw in reader:
            keyed = {(k or "").strip().lower(): (v or "").strip() for k, v in raw.items()}
            rows.append(_pick_from_keyed(keyed))
        return rows

    reader = csv.reader(stream, dialect=dialect)
    for cols in reader:
        item = _row_from_cells([_cell_str(c) for c in cols], has_header=False)
        if item:
            rows.append(item)
    return rows


def _iter_rows_from_xlsx(content: bytes) -> list[dict[str, str | None]]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    matrix: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = [_cell_str(c) for c in row]
        if any(cells):
            matrix.append(cells)
    wb.close()
    return _matrix_to_rows(matrix)


def _iter_rows_from_xls(content: bytes) -> list[dict[str, str | None]]:
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    matrix: list[list[str]] = []
    for r in range(sheet.nrows):
        cells = [_cell_str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
        if any(cells):
            matrix.append(cells)
    return _matrix_to_rows(matrix)


def _matrix_to_rows(matrix: list[list[str]]) -> list[dict[str, str | None]]:
    if not matrix:
        return []
    has_header = _headerish(matrix[0])
    rows: list[dict[str, str | None]] = []
    headers: list[str] | None = None
    start = 0
    if has_header:
        headers = [(c or "").strip().lower() for c in matrix[0]]
        start = 1
    for cells in matrix[start:]:
        item = _row_from_cells(cells, has_header=has_header, headers=headers)
        if item:
            rows.append(item)
    return rows


def _detect_kind(filename: str | None, content: bytes) -> str:
    name = (filename or "").lower().strip()
    if name.endswith(".xlsx"):
        return "xlsx"
    if name.endswith(".xls"):
        return "xls"
    if name.endswith(".csv") or name.endswith(".txt"):
        return "csv"
    # magic / fallback
    if content[:2] == b"PK":
        return "xlsx"
    if content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "xls"
    return "csv"


async def _insert_rows(db: AsyncSession, parsed: list[dict[str, str | None]]) -> ImportResponse:
    inserted = 0
    skipped = 0
    errors = 0

    for row in parsed:
        try:
            username = _norm_username(row.get("username") or "")
            user_id = _parse_user_id(str(row.get("user_id") or ""))

            if not username and not user_id:
                errors += 1
                continue

            stmt = insert(InviteTarget).values(username=username, user_id=user_id).prefix_with("OR IGNORE")
            res = await db.execute(stmt)
            if res.rowcount and res.rowcount > 0:
                inserted += 1
            else:
                skipped += 1
        except Exception:
            errors += 1

    return ImportResponse(inserted=inserted, skipped_duplicates=skipped, errors=errors)


async def import_targets_file(
    db: AsyncSession,
    content: bytes,
    filename: str | None = None,
) -> ImportResponse:
    kind = _detect_kind(filename, content)

    if kind == "xlsx":
        parsed = _iter_rows_from_xlsx(content)
    elif kind == "xls":
        parsed = _iter_rows_from_xls(content)
    else:
        text = content.decode("utf-8-sig", errors="replace")
        parsed = _iter_rows_from_text(text)
        if not parsed:
            for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
                line = line.strip()
                if not line or line.lower().startswith("username"):
                    continue
                if line.isdigit():
                    parsed.append({"username": None, "user_id": line})
                else:
                    parsed.append({"username": line, "user_id": None})

    return await _insert_rows(db, parsed)


# backward-compatible alias
async def import_targets_csv(db: AsyncSession, content: bytes) -> ImportResponse:
    return await import_targets_file(db, content, filename="targets.csv")
