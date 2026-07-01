"""Excel: лист 1 — карта клонов, лист 2 — схема работы."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
BATCH = ROOT / "forum_clones_batch.json"
OUT = ROOT / "VERDI_клоны_и_схема.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUB_FONT = Font(bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
POOL_FILL = {
    "verdi_7": PatternFill("solid", fgColor="E2EFDA"),
    "verdi_10": PatternFill("solid", fgColor="DDEBF7"),
    "verdi_13": PatternFill("solid", fgColor="FCE4D6"),
}
FREE_FILL = PatternFill("solid", fgColor="F2F2F2")


def style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[letter].width = min(width, max_width)


def sheet_map(wb: Workbook) -> None:
    batch = json.loads(BATCH.read_text(encoding="utf-8"))
    ws = wb.active
    ws.title = "Карта клонов"

    ws["A1"] = "VERDI — карта исходников и клонов"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")

    headers = [
        "№ клона",
        "Роль клона",
        "Название клона",
        "Ссылка клона",
        "Пул",
        "Исходник (название)",
        "ID исходника",
        "Статус",
        "Техпул",
    ]
    row = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))

    for clone in sorted(batch["clones"], key=lambda x: x["index"]):
        row += 1
        pool = clone.get("pool", "")
        status = "Привязан" if clone.get("source_chat_id") else "Свободен"
        tech = "tech_1 … tech_10" if pool else "—"
        values = [
            clone["index"],
            clone.get("role", "—"),
            clone.get("mirror_title", ""),
            clone.get("public_link", ""),
            pool or "—",
            clone.get("source_title", "—"),
            clone.get("source_chat_id", "—"),
            status,
            tech,
        ]
        fill = POOL_FILL.get(pool, FREE_FILL if status == "Свободен" else None)
        for i, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=val)
            cell.border = BORDER
            cell.alignment = WRAP
            if fill:
                cell.fill = fill

    row += 2
    ws.cell(row=row, column=1, value="Сводка пулов").font = SUB_FONT
    pools = batch.get("pools", {})
    for pool_name, pdata in pools.items():
        row += 1
        ws.cell(row=row, column=1, value=pool_name).font = Font(bold=True)
        for r in pdata.get("routes", []):
            row += 1
            ws.cell(
                row=row,
                column=2,
                value=f"{r['source_title']}  →  {r['mirror_public_link']}",
            ).alignment = WRAP

    autosize(ws)


def sheet_scheme(wb: Workbook) -> None:
    ws = wb.create_sheet("Как работает клонер")
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 28

    ws["B2"] = "Схема клонирования (пул VERDI 7 — пример)"
    ws["B2"].font = TITLE_FONT
    ws.merge_cells("B2:F2")

    blocks = [
        ("B4", "F4", "ИСТОЧНИКИ\n(не трогаем техаккаунтами)", "4472C4"),
        ("B5", "C5", "VERDI 7\n«отчеты»", "D9E1F2"),
        ("D5", "E5", "VERDI 7\n«задания»", "D9E1F2"),
        ("B8", "F8", "СЛУШАТЕЛЬ\n@andf1n (listener_main)\nмодератор в источниках", "7030A0"),
        ("B11", "F11", "ПУЛ ТЕХАККАУНТОВ\ntech_1 … tech_10\nкаждый участник → свой tech", "548235"),
        ("B14", "C14", "КЛОН\nmultiverdichat1\nотчёты", "E2EFDA"),
        ("E14", "F14", "КЛОН\nmultiverdichat4\nзадание", "E2EFDA"),
    ]

    for start, end, text, color in blocks:
        ws.merge_cells(f"{start}:{end}")
        cell = ws[start]
        cell.value = text
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color="FFFFFF" if color in ("4472C4", "7030A0", "548235") else "000000")
        cell.alignment = CENTER
        cell.border = BORDER

    arrows = ["D6", "D9", "D12", "C12", "E12"]
    for addr in arrows:
        c = ws[addr]
        c.value = "▼" if addr in ("D6", "D9", "D12") else "↙"
        c.alignment = CENTER
        c.font = Font(size=16, bold=True)

    ws["B17"] = "Пошагово"
    ws["B17"].font = SUB_FONT
    steps = [
        "1. В исходном чате участник пишет сообщение.",
        "2. listener_main (@andf1n) видит сообщение — он модератор, техаккаунты в источник НЕ добавляются.",
        "3. Система выбирает техаккаунт из пула (первый раз — свободный, дальше тот же за этим user_id).",
        "4. Техаккаунт копирует имя и аватар отправителя.",
        "5. Техаккаунт публикует текст/медиа в клон-чат (multiverdichat1 или multiverdichat4).",
        "6. Участники клона видят сообщение как от «живого» профиля.",
    ]
    for i, step in enumerate(steps, 18):
        ws.cell(row=i, column=2, value=step).alignment = WRAP
        ws.merge_cells(f"B{i}:F{i}")

    ws["B25"] = "Запуск пула VERDI 7"
    ws["B25"].font = SUB_FONT
    cmds = [
        "python prepare_mirrors.py --config multi_config.verdi7.json",
        "python run_pool.py --config multi_config.verdi7.json",
    ]
    for i, cmd in enumerate(cmds, 26):
        c = ws.cell(row=i, column=2, value=cmd)
        c.font = Font(name="Consolas", size=10)
        ws.merge_cells(f"B{i}:F{i}")

    note_row = 29
    ws.cell(
        row=note_row,
        column=2,
        value="Файл привязок: sender_bindings_verdi7.json (кто за каким tech закреплён)",
    ).alignment = WRAP
    ws.merge_cells(f"B{note_row}:F{note_row}")


def main() -> None:
    wb = Workbook()
    sheet_map(wb)
    sheet_scheme(wb)
    wb.save(OUT)
    print(f"Сохранено: {OUT}")


if __name__ == "__main__":
    main()
